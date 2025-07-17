from tkinter import *
from tkinter import ttk, messagebox

import numpy as np
import pandas as pd
import openpyxl as xl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt

class CalculadoraLucro():
    def __init__(self):
        self.root = Tk()
        self.root.title("Quitanda do José")
        self.root.geometry("1100x600")
        self.root.resizable(False, False)

        self.color_green = "#2ca063"
        self.color_greendois = "#0aaf57"
        self.color_yellow = "#e6a04b"
        self.color_white = "#f8f8f8"

        self.containers()
        self.itens_container01()
        self.itens_container02()
        self.itens_container03()
        self.root.mainloop()

    def containers(self):
        self.fr_container01 = Frame(
            self.root,
            width=1200,
            height=30,
            bg=self.color_greendois,
            bd=2,
            relief='solid'
        )
        self.fr_container02 = Frame(
            self.root,
            width=1050,
            height=1000,
            bg=self.color_green,
            bd=2,
            relief='solid'
        )
        self.fr_container03 = Frame(
            self.root,
            width=1100,
            height=370,
            bg='red',
            bd=2,
            relief='solid'
        )

        self.fr_container01.propagate(0)
        self.fr_container02.propagate(0)
        self.fr_container03.propagate(0)
        self.fr_container01.pack()
        self.fr_container02.pack()
        self.fr_container03.pack()

    def itens_container01(self):
        self.lb_title = Label(
            self.fr_container01,
            text='Calculadora de Lucros',
            font='Verdana 20',
            bg=self.fr_container01.cget('bg')
        )
        self.lb_title.pack()

    def itens_container02(self):
        self.fr_subcontainer01 = Frame(
            self.fr_container02,
            bg='white',
            highlightthickness=1,
            highlightcolor='gray'
        )
        self.fr_subcontainer02 = Frame(
            self.fr_container02,
            bg='white',
            highlightthickness=1,
            highlightcolor='gray'
        )
        self.fr_subcontainer03 = Frame(
            self.fr_container02,
            bg='white',
            highlightthickness=1,
            highlightcolor='gray'
        )

        self.fr_container_btn = Frame(
            self.fr_subcontainer01,
            bg=self.fr_subcontainer01.cget('bg')
        )

        # secao 01
        self.lb_title_section01 = Label(
            self.fr_subcontainer01,
            text='Cadastro de produto',
            font='Verdana 15',
            bg=self.fr_subcontainer01.cget('bg')
        )

        # Nome do produto
        self.lb_nome_produto = Label(
            self.fr_subcontainer01,
            text='Nome do produto',
            font='Verdana',
            bg=self.fr_subcontainer01.cget('bg')
        )
        self.en_nome_produto = Entry(
            self.fr_subcontainer01,
            font='Verdana',
            highlightthickness=1,
            highlightbackground='gray',
            highlightcolor=self.color_green,
            bg='white'
        )

        # Quantidade do produto
        self.lb_qtd_produto = Label(
            self.fr_subcontainer01,
            text='Quantidade do produto',
            font='Verdana',
            bg=self.fr_subcontainer01.cget('bg')
        )
        self.en_qtd_produto = Entry(
            self.fr_subcontainer01,
            font='Verdana',
            highlightthickness=1,
            highlightbackground='gray',
            highlightcolor=self.color_green,
            bg='white'
        )

        # Preço de compra
        self.lb_preco_compra = Label(
            self.fr_subcontainer01,
            text='Preço de compra',
            font='Verdana',
            bg=self.fr_subcontainer01.cget('bg')
        )
        self.en_preco_compra = Entry(
            self.fr_subcontainer01,
            font='Verdana',
            highlightthickness=1,
            highlightbackground='gray',
            highlightcolor=self.color_green,
            bg='white'
        )

        # Preço de venda
        self.lb_preco_venda = Label(
            self.fr_subcontainer01,
            text='Preço de venda',
            font='Verdana',
            bg=self.fr_subcontainer01.cget('bg')
        )
        self.en_preco_venda = Entry(
            self.fr_subcontainer01,
            font='Verdana',
            highlightthickness=1,
            highlightbackground='gray',
            highlightcolor=self.color_green,
            bg='white'
        )

        # Custo frete
        self.lb_custo_frete = Label(
            self.fr_subcontainer01,
            text='Custo frete',
            font='Verdana',
            bg=self.fr_subcontainer01.cget('bg')
        )
        self.en_custo_frete = Entry(
            self.fr_subcontainer01,
            font='Verdana',
            highlightthickness=1,
            highlightbackground='gray',
            highlightcolor=self.color_green,
            bg='white'
        )

        # Custo adicional
        self.lb_custo_adicional = Label(
            self.fr_subcontainer01,
            text='Custo adicional',
            font='Verdana',
            bg=self.fr_subcontainer01.cget('bg')
        )
        self.en_custo_adicional = Entry(
            self.fr_subcontainer01,
            font='Verdana',
            highlightthickness=1,
            highlightbackground='gray',
            highlightcolor=self.color_green,
            bg='white'
        )

        # Botões
        self.btn_calcular = Button(
            self.fr_container_btn,
            text='Calcular',
            fg='white',
            bg='#0097b2',
            font='Verdana',
            command=self.calcular
        )

        self.btn_salvar = Button(
            self.fr_container_btn,
            text='Salvar',
            fg='white',
            bg=self.color_green,
            font = 'Verdana',
            command=self.salvar_registro
        )

        self.btn_deletar = Button(
            self.fr_container_btn,
            text='Deletar',
            fg='white',
            bg='#ff3131',
            command=self.excluir_registro
        )

        # secao 02
        self.lb_title_section02 = Label(
            self.fr_subcontainer02,
            text='Operações',
            font='Verdana 15',
            bg=self.fr_subcontainer02.cget('bg')
        )

        self.lb_texto_resultado_operacao = Label(
            self.fr_subcontainer02,
            text='O resultado aparecerá aqui assim que calcular um produto',
            font='Verdana',
            wraplength=250,
            justify=LEFT,
            bg=self.fr_subcontainer02.cget('bg')
        )

        self.lb_title_lucro_liquido = Label(
            self.fr_subcontainer02,
            text='Lucro liquido',
            font='Verdana',
            bg=self.fr_subcontainer02.cget('bg')
        )

        self.lb_resultado_lucro_liquido = Label(
            self.fr_subcontainer02,
            text='R$000,00',
            font='Verdana',
            width=20,
            padx=100,
            pady=5,
            highlightthickness=1,
            highlightbackground='gray',
            bg=self.fr_subcontainer02.cget('bg')
        )

        self.lb_title_margem_lucro = Label(
            self.fr_subcontainer02,
            text='Margem de Lucro',
            font='Verdana',
            bg=self.fr_subcontainer02.cget('bg')
        )

        self.lb_resultado_margem_lucro = Label(
            self.fr_subcontainer02,
            text='0,00%',
            font='Verdana',
            width=20,
            padx=100,
            pady=5,
            highlightthickness=1,
            highlightbackground='gray',
            bg=self.fr_subcontainer02.cget('bg')
        )

        # Alocando os containers
        self.fr_subcontainer01.grid(row=0, column=0, padx=10, pady=7, sticky=N)
        self.fr_subcontainer02.grid(row=0, column=1, padx=10, pady=7, sticky=N)
        self.fr_subcontainer03.grid(row=0, column=2, padx=10, pady=7, sticky=N)

        # Posicionando elementos subcontainer01
        self.lb_title_section01.grid(row=0, columnspan=2, pady=10)
        self.lb_nome_produto.grid(row=1, column=0, sticky=W, padx=5, pady=5)
        self.en_nome_produto.grid(row=1, column=1, sticky=W, padx=5)
        self.lb_qtd_produto.grid(row=2, column=0, sticky=W, padx=5, pady=5)
        self.en_qtd_produto.grid(row=2, column=1, sticky=W, padx=5)
        self.lb_preco_compra.grid(row=3, column=0, sticky=W, padx=5, pady=5)
        self.en_preco_compra.grid(row=3, column=1, sticky=W, padx=5)
        self.lb_preco_venda.grid(row=4, column=0, sticky=W, padx=5, pady=5)
        self.en_preco_venda.grid(row=4, column=1, sticky=W, padx=5)
        self.lb_custo_frete.grid(row=5, column=0, sticky=W, padx=5, pady=5)
        self.en_custo_frete.grid(row=5, column=1, sticky=W, padx=5)
        self.lb_custo_adicional.grid(row=6, column=0, sticky=W, padx=5, pady=5)
        self.en_custo_adicional.grid(row=6, column=1, sticky=W, padx=5)

        self.fr_container_btn.grid(row=7, columnspan=2, padx=10, sticky=W)
        self.btn_calcular.grid(row=0, column=0, sticky=W, padx=5)
        self.btn_salvar.grid(row=0, column=1, sticky=W, padx=5)
        self.btn_deletar.grid(row=0, column=2, sticky=W, padx=5)

        # Posicionando elementos do subcontainer02
        self.lb_title_section02.grid(row=0, column=0, padx=10)
        self.lb_texto_resultado_operacao.grid(row=1, column=0, padx=30)
        self.lb_title_lucro_liquido.grid(row=2, column=0)
        self.lb_resultado_lucro_liquido.grid(row=3, column=0, padx=5)
        self.lb_title_margem_lucro.grid(row=4, column=0)
        self.lb_resultado_margem_lucro.grid(row=5, column=0, padx=5, pady=(0, 60))

        self.grafico()
        
    def itens_container03(self):
       colunas_tabela = [
        'Nome do Produto', 'Preço de Compra(R$)', 'Preco de Venda(R$)', 'Qtd',
        'Custos adicionais(R$)', 'Custo médio do frete(R$)', 'Custo total(R$)',
        'Lucro Liquido(R$)', 'Margem de Lucro(%)'
    ]

    # Frame interno para organizar a Treeview com scrollbars
       frame_tabela = Frame(self.fr_container03)
       frame_tabela.pack(fill=BOTH, expand=True)

    # Scrollbars criadas no MESMO container da treeview
       v_scroll = Scrollbar(frame_tabela, orient=VERTICAL)
       h_scroll = Scrollbar(frame_tabela, orient=HORIZONTAL)

       self.treeview = ttk.Treeview(
          frame_tabela,
          columns=colunas_tabela,
          show='headings',
          yscrollcommand=v_scroll.set,
          xscrollcommand=h_scroll.set
    )

       for col in colunas_tabela:
           self.treeview.heading(col, text=col)
           self.treeview.column(col, width=180, anchor=CENTER)

    # Configura scrolls
       v_scroll.config(command=self.treeview.yview)
       h_scroll.config(command=self.treeview.xview)

    # Posicionamento via grid
       self.treeview.grid(row=0, column=0, sticky='nsew')
       v_scroll.grid(row=0, column=1, sticky='ns')
       h_scroll.grid(row=1, column=0, sticky='ew')

    # Expansão automática
       frame_tabela.grid_rowconfigure(0, weight=1)
       frame_tabela.grid_columnconfigure(0, weight=1)

       self.atualizar_tabela()

    # função para atualizar a tabela quando houver alterações
    def atualizar_tabela(self):
        self.grafico()
        df_list = self.obter_dados_tabela('planilha.xlsx')

        for i in self.treeview.get_children():
            self.treeview.delete(i)

        for item in df_list:
            self.treeview.insert('', 'end', values=item)
        self.grafico()

    def obter_dados_tabela(self, nome_planilha):
        wb = xl.load_workbook(nome_planilha)

        sheet = wb.active
        dados = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            dados.append(row)

        return dados
    #Função para salvar dados na planilha
    def salvar_registro(self):
        if self.valiar_entrys() == True:
            nome_produto = self.en_nome_produto.get()
            qtd_produto = float(self.en_qtd_produto.get())
            preco_venda = float(self.en_preco_venda.get())
            preco_compra = float(self.en_preco_compra.get())
            custo_adicional = float(self.en_custo_adicional.get())
            custo_frete = float(self.en_custo_frete.get())

            lucro_liquido = preco_venda - preco_compra - custo_adicional - custo_frete

            custo_geral = (preco_compra+ custo_adicional + custo_frete) * qtd_produto

            margem_lucro = (lucro_liquido/preco_venda) * 100
           #Carregar planilha existente ou criar uma nova
            try:
                wb = xl.load_workbook('planilha.xlsx')
                sheet = wb.active
            except:
                wb = xl.Workbook()
                sheet = wb.active
                sheet.title('Plan1')

            #Adicionando cabeçalhos
                sheet['A1'] = 'Nome do produto'
                sheet['B1'] = 'Preco da compra'
                sheet['C1'] = 'Preco da venda'
                sheet['D1'] = 'Quantidade'
                sheet['E1'] = 'Custos Adicionais'
                sheet['F1'] = 'Custo do Frete'
                sheet['G1'] = 'Custo Total'
                sheet['H1'] = 'Lucro Liquido'
                sheet['I1'] = 'Margem de Lucro (%)'

            #Adicionando valores na minha planilha
            row = sheet.max_row + 1
            sheet[f'A{row}'] = nome_produto
            sheet[f'B{row}'] = preco_compra
            sheet[f'C{row}'] = preco_venda
            sheet[f'D{row}'] = qtd_produto
            sheet[f'E{row}'] = custo_adicional
            sheet[f'F{row}'] = custo_frete
            sheet[f'G{row}'] = custo_geral
            sheet[f'H{row}'] = lucro_liquido
            sheet[f'I{row}'] = margem_lucro


            wb.save('planilha.xlsx')
            self.atualizar_tabela()
            self.resetar_entrys()

        else:
            messagebox.showinfo('Campo vazio, existe algum campo obrigatorio vazio!')

    #Calcular registros
    def calcular(self):
        if self.valiar_entrys() == True:
             if self.valiar_entrys() == True:
                nome_produto = self.en_nome_produto.get()
                qtd_produto = float(self.en_qtd_produto.get())
                preco_venda = float(self.en_preco_venda.get())
                preco_compra = float(self.en_preco_compra.get())
                custo_adicional = float(self.en_custo_adicional.get())
                custo_frete = float(self.en_custo_frete.get())

                custo_geral = (preco_compra * qtd_produto) + custo_adicional + custo_frete
                lucro_liquido = preco_venda - custo_geral
                margem_lucro = (lucro_liquido / preco_venda) * 100

                self.lb_texto_resultado_operacao['text'] = 'O lucro do produto {} é de R${:.2f} e a margem de lucro é de {:.2f}%'.format(nome_produto, lucro_liquido, margem_lucro)

                self.lb_resultado_lucro_liquido['text'] = 'R${:.2f}'.format(lucro_liquido)
                self.lb_resultado_margem_lucro['text'] = 'R${:.2f}%'.format(margem_lucro)

        else: 
            messagebox.showinfo('Campo vazio, existe algum campo obrigatorio vazio!')

    def excluir_registro(self):
        try:
            dado_selecionado = self.treeview.focus()
            treev_dicionario = self.treeview.item(dado_selecionado)
            treev_lista = treev_dicionario['values']
            valor = treev_lista[0]

            wb = xl.load_workbook('planilha.xlsx')
            sheet = wb.active
            contador = 2

            for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 1, values_only = True):
                if str(row[0]) == valor:
                    linha = contador
                    sheet.delete_rows(linha)
                    messagebox.showinfo('Exclusão', f'{valor} foi deletado com sucesso!')
                    break

                contador += 1

            wb.save('planilha.xlsx')
            self.atualizar_tabela()
        except: 
            messagebox.showinfo('Campos vazios',f'Selecione um produto para deletar')

    def valiar_entrys(self):
        campos = [
            self.en_nome_produto.get(),
            self.en_qtd_produto.get(),
            self.en_preco_compra.get(),
            self.en_preco_venda.get(),
            self.en_custo_adicional.get(),
            self.en_custo_frete.get()
        ]
        if all(campo.strip() for campo in campos):
            return True
        else: 
            return False
        #quero que todos os itens atribuidos dentro de campo tiver algum vazio me retorne false, caso ao contario retorne true.

    def resetar_entrys(self):
        self.en_nome_produto.delete(0, 'end'),
        self.en_qtd_produto.delete(0, 'end'),
        self.en_preco_compra.delete(0, 'end'),
        self.en_preco_venda.delete(0, 'end'),
        self.en_custo_adicional.delete(0, 'end'),
        self.en_custo_frete.delete(0, 'end'),

    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt

    def grafico(self):
        self.lib_title_section03 = Label(
            self.fr_subcontainer03,
            text='Estatistica dos produtos salvos',
            font='Verdana 15',
            bg=self.fr_subcontainer03.cget('bg')
        )

        lista_nomes = ['Total custos', 'Lucro liquido total', 'margem de lucro total']
        lista_valores = self.estatistica()
        simbol = ['R$', 'R$', '%']

        figura = plt.Figure(figsize=(4.5, 3.3), dpi=70)
        ax = figura.add_subplot(111)

        ax.bar(
            lista_nomes,
            lista_valores,
            color = self.color_green,
            width = 0.5
        )

        c = 0
        for i in ax.patches:
            if c == 2:
                ax.text(i.get_x()-.001, i.get_height()+.5, str("{:,.0f}".format(lista_valores[c]) +simbol[c]), fontsize = 17, fontstyle = 'italic', verticalalignment = 'bottom', color='dimgray')
            else:
               ax.text(i.get_x()-.001, i.get_height()+.5, str(simbol[c]+"{:,.0f}".format(lista_valores[c])),fontsize = 17, fontstyle = 'italic', verticalalignment = 'bottom', color='dimgray')
            c += 1

        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)

        canva = FigureCanvasTkAgg(
            figura,
            self.fr_subcontainer03
        )

        self.lib_title_section03.grid(row=0, column=0, pady=(5,5))
        canva.get_tk_widget().grid(row=1, column=0, pady=(5,80))


    def estatistica(self):
        wb = xl.load_workbook('planilha.xlsx')
        sheet = wb.active

        primeira_linha = 2
        ultima_linha = sheet.max_row

        custo_total = 0
        lucro_total = 0

        for linha in range(primeira_linha, ultima_linha + 1):
            custo_total += sheet.cell(row=linha, column=7).value or 0
            lucro_total += sheet.cell(row=linha, column=8).value or 0

        preco_venda_total = sum(
            sheet.cell(row=linha, column=3).value or 0
            for linha in range(primeira_linha, ultima_linha + 1)
    )

        try:
            margem_lucro_total = (lucro_total / preco_venda_total) * 100
        except ZeroDivisionError:
            margem_lucro_total = 0

        custo_total = round(custo_total, 2)
        lucro_total = round(lucro_total, 2)
        margem_lucro_total = round(margem_lucro_total, 2)

        return [custo_total, lucro_total, margem_lucro_total]

CalculadoraLucro()
